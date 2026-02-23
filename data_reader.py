# -*- coding: utf-8 -*-
"""
PPT报告生成器 - 数据读取模块
从Excel文件中读取并解析CCTV-17频道收视数据
"""
import datetime
import openpyxl
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple


@dataclass
class MarketShareData:
    """市场份额数据"""
    period_label: str = ""           # 前推期间标签
    current_date_label: str = ""      # 当日标签
    # CCTV-17
    cctv17_period_rating: float = 0   # 前推期间收视率
    cctv17_period_share: float = 0    # 前推期间市场份额
    cctv17_period_reach: float = 0    # 前推期间到达率
    cctv17_period_loyalty: float = 0  # 前推期间忠实度
    cctv17_current_rating: float = 0  # 当日收视率
    cctv17_current_share: float = 0   # 当日市场份额
    cctv17_current_reach: float = 0   # 当日到达率
    cctv17_current_loyalty: float = 0 # 当日忠实度
    # 所有频道
    all_period_rating: float = 0
    all_period_share: float = 0
    all_period_reach: float = 0
    all_period_loyalty: float = 0
    all_current_rating: float = 0
    all_current_share: float = 0
    all_current_reach: float = 0
    all_current_loyalty: float = 0
    # 变化值（百分比）
    rating_change: float = 0
    share_change: float = 0
    reach_change: float = 0
    loyalty_change: float = 0
    # 前一日数据（如有）
    has_prev_day: bool = False
    prev_date_label: str = ""          # 前一日日期标签（如 "2月14日"）
    cctv17_prev_rating: float = 0
    cctv17_prev_share: float = 0
    cctv17_prev_reach: float = 0
    cctv17_prev_loyalty: float = 0
    all_prev_rating: float = 0
    all_prev_share: float = 0
    all_prev_reach: float = 0
    all_prev_loyalty: float = 0


@dataclass
class DramaData:
    """电视剧与非电视剧数据"""
    period_label: str = ""
    current_label: str = ""
    drama_period_rating: float = 0
    drama_period_share: float = 0
    drama_current_rating: float = 0
    drama_current_share: float = 0
    non_drama_period_rating: float = 0
    non_drama_period_share: float = 0
    non_drama_current_rating: float = 0
    non_drama_current_share: float = 0


@dataclass
class ChannelRankItem:
    """频道排名项"""
    rank: int = 0
    name: str = ""
    short_name: str = ""
    prev_share: float = 0
    current_share: float = 0
    change: float = 0


@dataclass
class ProgramItem:
    """节目项(串单)"""
    name: str = ""
    episode: str = ""
    date: str = ""
    weekday: str = ""
    start_time: str = ""
    duration: int = 0
    end_time: str = ""
    category: str = ""
    subcategory: str = ""
    market_share: float = 0
    rating: float = 0


@dataclass
class MinuteData:
    """分钟数据"""
    time_str: str = ""
    period_rating: float = 0
    current_rating: float = 0
    period_share: float = 0
    current_share: float = 0


@dataclass
class TimeSlotData:
    """分时段收视率数据"""
    slot_name: str = ""
    period_rating: float = 0
    current_rating: float = 0
    period_share: float = 0
    current_share: float = 0
    rating_change: float = 0
    share_change: float = 0


@dataclass
class TheaterData:
    """三大剧场数据"""
    name: str = ""
    time_slot: str = ""
    info: str = ""
    period_rating: float = 0
    period_share: float = 0
    current_rating: float = 0
    current_share: float = 0
    rating_change: float = 0
    share_change: float = 0


@dataclass
class PremiereProgram:
    """首播节目对比"""
    name: str = ""
    time_slot: str = ""
    rating: float = 0
    share: float = 0
    prev_day_rating: float = 0
    prev_day_share: float = 0
    period_rating: float = 0
    period_share: float = 0
    rating_change: float = 0
    share_change: float = 0
    # 忠实度(运行时计算)
    loyalty: float = 0
    loyalty_change: float = 0
    audience: float = 0  # 四岁以上观众规模(万人)


@dataclass
class AudienceItem:
    """观众规模项"""
    category: str = ""
    period_value: float = 0
    current_value: float = 0
    change: float = 0


@dataclass
class ReportData:
    """完整报告数据"""
    # 元数据
    report_date: str = ""         # 报告日期 如 "2026年2月10日"
    report_date_short: str = ""   # 如 "2月10日"
    period_label: str = ""        # 前推期间 如 "2026/1/11-2026/2/9"

    # 基础数据(6个基本sheet)
    market_share: MarketShareData = field(default_factory=MarketShareData)
    drama: DramaData = field(default_factory=DramaData)
    org_ranking: List[ChannelRankItem] = field(default_factory=list)     # 台组内排名
    channel_ranking: List[ChannelRankItem] = field(default_factory=list) # 上星频道排名
    programs: List[ProgramItem] = field(default_factory=list)            # 串单
    minutes: List[MinuteData] = field(default_factory=list)              # 分分钟数据

    # 扩展数据(日报sheets)
    has_daily_report: bool = False
    time_slots: List[TimeSlotData] = field(default_factory=list)         # 日报-分时段
    theaters: List[TheaterData] = field(default_factory=list)            # 日报-三大剧场
    premiere_programs: List[PremiereProgram] = field(default_factory=list)# 日报-首播节目对比
    premiere_audience: Dict = field(default_factory=dict)                 # 日报-首播分类观众
    channel_audience: List[AudienceItem] = field(default_factory=list)   # 日报-频道分类观众
    drama_audience: Dict = field(default_factory=dict)                    # 日报-电视剧观众规模

    # 排名信息(计算字段)
    org_rank: int = 0
    org_rank_change: int = 0
    channel_rank: int = 0
    channel_rank_change: int = 0


def _safe_float(val, default=0.0):
    """安全转换为浮点数"""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_str(val):
    """安全转换为字符串"""
    if val is None:
        return ""
    return str(val).strip()


def _time_to_str(val):
    """将时间值转换为字符串"""
    if val is None:
        return ""
    if isinstance(val, datetime.time):
        return val.strftime('%H:%M')
    if isinstance(val, datetime.timedelta):
        total_seconds = int(val.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f'{hours:02d}:{minutes:02d}'
    return str(val).strip()


def _date_to_str(val):
    """将日期值转为字符串"""
    if val is None:
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime('%Y/%m/%d')
    return str(val).strip()


def _extract_date_from_label(label: str) -> str:
    """从日期标签中提取日期，如 '2026/2/10' -> '2月10日'"""
    if not label:
        return ""
    parts = label.replace('-', '/').split('/')
    if len(parts) >= 3:
        try:
            month = int(parts[1])
            day = int(parts[2].split(' ')[0])
            return f'{month}月{day}日'
        except:
            pass
    return label


def read_excel_data(filepath: str) -> ReportData:
    """
    读取Excel数据并返回结构化的ReportData

    支持两种格式:
    1. 基础格式(6个sheet): 市场份额, 电视剧与非电视剧, 上星频道排名, 台组内排名, 分分钟, 串单
    2. 完整格式(12个sheet): 基础6个 + 日报-分时段收视率, 日报-三大剧场, 日报-电视剧观众规模,
       日报-首播节目对比, 日报-首播节目分类观众触达, 日报-频道分类观众规模
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = ReportData()

    sheet_names = wb.sheetnames

    # 读取市场份额
    if '市场份额' in sheet_names:
        _read_market_share(wb['市场份额'], data)
        # 从市场份额H-W列读取分类观众规模（优先级高于日报-频道分类观众规模）
        _read_audience_from_market_share(wb['市场份额'], data)

    # 读取电视剧与非电视剧
    if '电视剧与非电视剧' in sheet_names:
        _read_drama(wb['电视剧与非电视剧'], data)

    # 读取台组内排名
    if '台组内排名' in sheet_names:
        _read_org_ranking(wb['台组内排名'], data)

    # 读取上星频道排名
    if '上星频道排名' in sheet_names:
        _read_channel_ranking(wb['上星频道排名'], data)

    # 读取串单
    if '串单' in sheet_names:
        _read_programs(wb['串单'], data)
        # 从串单读取首播节目数据（16:30-24:00 非电视剧）
        _read_premiere_from_schedule(wb['串单'], data)

    # 读取分分钟
    if '分分钟' in sheet_names:
        _read_minutes(wb['分分钟'], data)

    # 读取日报数据
    daily_sheets = ['日报-分时段收视率', '日报-三大剧场', '日报-电视剧观众规模',
                    '日报-首播节目对比', '日报-首播节目分类观众触达', '日报-频道分类观众规模']
    if any(s in sheet_names for s in daily_sheets):
        data.has_daily_report = True
        if '日报-分时段收视率' in sheet_names:
            _read_time_slots(wb['日报-分时段收视率'], data)
        if '日报-三大剧场' in sheet_names:
            _read_theaters(wb['日报-三大剧场'], data)
        if '日报-首播节目对比' in sheet_names:
            # 串单已提供首播节目数据时跳过（串单数据更准确、时效性更好）
            if not data.premiere_programs:
                _read_premiere_programs(wb['日报-首播节目对比'], data)
        if '日报-首播节目分类观众触达' in sheet_names:
            _read_premiere_audience(wb['日报-首播节目分类观众触达'], data)
        if '日报-频道分类观众规模' in sheet_names:
            # 市场份额H-W列已提供分类观众数据时跳过（数据更准确）
            if not data.channel_audience:
                _read_channel_audience(wb['日报-频道分类观众规模'], data)
        if '日报-电视剧观众规模' in sheet_names:
            _read_drama_audience(wb['日报-电视剧观众规模'], data)

    wb.close()

    # 计算排名
    _compute_rankings(data)

    # 计算首播节目忠实度
    _compute_premiere_loyalty(data)

    return data


def _read_market_share(ws, data: ReportData):
    """读取市场份额sheet"""
    ms = data.market_share
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 3:
        return

    # 判断行数确定数据格式
    # 格式1 (7行,含变化列): CCTV-17前推/前一日/当日, 所有频道前推/前一日/当日
    # 格式2 (5行): CCTV-17前推/当日, 所有频道前推/当日

    # 找到CCTV-17的行和所有频道的行
    cctv17_rows = []
    all_rows = []
    for r in rows[1:]:  # 跳过表头
        if r[0] and '农业农村' in str(r[0]):
            cctv17_rows.append(r)
        elif r[0] and '所有频道' in str(r[0]):
            all_rows.append(r)

    if not cctv17_rows:
        return

    # 前推期间数据
    ms.period_label = _safe_str(cctv17_rows[0][1])
    ms.cctv17_period_rating = _safe_float(cctv17_rows[0][2])
    ms.cctv17_period_share = _safe_float(cctv17_rows[0][3])
    ms.cctv17_period_reach = _safe_float(cctv17_rows[0][4])
    ms.cctv17_period_loyalty = _safe_float(cctv17_rows[0][5])
    data.period_label = ms.period_label

    # 当日数据 (最后一行CCTV-17)
    current_row = cctv17_rows[-1]
    ms.current_date_label = _safe_str(current_row[1])
    ms.cctv17_current_rating = _safe_float(current_row[2])
    ms.cctv17_current_share = _safe_float(current_row[3])
    ms.cctv17_current_reach = _safe_float(current_row[4])
    ms.cctv17_current_loyalty = _safe_float(current_row[5])

    # 提取报告日期
    date_str = ms.current_date_label
    data.report_date_short = _extract_date_from_label(date_str)
    if '/' in date_str:
        parts = date_str.split('/')
        try:
            data.report_date = f'{parts[0]}年{int(parts[1])}月{int(parts[2])}日'
        except:
            data.report_date = date_str

    # 前一日数据 (如果有3行CCTV-17数据)
    if len(cctv17_rows) >= 3:
        ms.has_prev_day = True
        prev = cctv17_rows[1]
        ms.cctv17_prev_rating = _safe_float(prev[2])
        ms.cctv17_prev_share = _safe_float(prev[3])
        ms.cctv17_prev_reach = _safe_float(prev[4])
        ms.cctv17_prev_loyalty = _safe_float(prev[5])
        # 前一日日期标签
        prev_label = _safe_str(prev[1])
        ms.prev_date_label = _extract_date_from_label(prev_label)

    # 所有频道数据
    if all_rows:
        ms.all_period_rating = _safe_float(all_rows[0][2])
        ms.all_period_share = _safe_float(all_rows[0][3])
        ms.all_period_reach = _safe_float(all_rows[0][4])
        ms.all_period_loyalty = _safe_float(all_rows[0][5])
        # 所有频道前一日数据
        if len(all_rows) >= 3:
            ms.all_prev_rating = _safe_float(all_rows[1][2])
            ms.all_prev_share = _safe_float(all_rows[1][3])
            ms.all_prev_reach = _safe_float(all_rows[1][4])
            ms.all_prev_loyalty = _safe_float(all_rows[1][5])
        ms.all_current_rating = _safe_float(all_rows[-1][2])
        ms.all_current_share = _safe_float(all_rows[-1][3])
        ms.all_current_reach = _safe_float(all_rows[-1][4])
        ms.all_current_loyalty = _safe_float(all_rows[-1][5])

    # 计算变化百分比
    if ms.cctv17_period_share > 0:
        ms.share_change = round((ms.cctv17_current_share - ms.cctv17_period_share) / ms.cctv17_period_share * 100)
    if ms.cctv17_period_rating > 0:
        ms.rating_change = round((ms.cctv17_current_rating - ms.cctv17_period_rating) / ms.cctv17_period_rating * 100)
    if ms.cctv17_period_reach > 0:
        ms.reach_change = round((ms.cctv17_current_reach - ms.cctv17_period_reach) / ms.cctv17_period_reach * 100)
    if ms.cctv17_period_loyalty > 0:
        ms.loyalty_change = round((ms.cctv17_current_loyalty - ms.cctv17_period_loyalty) / ms.cctv17_period_loyalty * 100)

    # 如有变化列(7列以上), 直接使用
    # 注意：仅当表头明确标有"变化"字样时才覆盖计算值
    # 当前 Excel 格式的第7列起为"观众规模"等数据，不是变化列
    if len(rows) > 1 and len(rows[0]) >= 10:
        header_6 = _safe_str(rows[0][6]).lower() if rows[0][6] else ''
        if '变化' in header_6 or 'change' in header_6:
            last_cctv = cctv17_rows[-1]
            if len(last_cctv) >= 10 and last_cctv[6] is not None:
                ms.rating_change = _safe_float(last_cctv[6])
                ms.share_change = _safe_float(last_cctv[7])
                ms.reach_change = _safe_float(last_cctv[8])
                ms.loyalty_change = _safe_float(last_cctv[9])


def _read_audience_from_market_share(ws, data: ReportData):
    """从市场份额sheet H-W列读取分类观众规模

    列映射 (0-indexed):
      G(6)=观众规模(总体), H(7)=男, I(8)=女,
      J(9)=4-14岁, K(10)=15-24岁, L(11)=25-34岁, M(12)=35-44岁,
      N(13)=45-54岁, O(14)=55-64岁, P(15)=65岁以上,
      Q(16)=未受过正规教育, R(17)=小学, S(18)=初中, T(19)=高中, U(20)=大学以上,
      V(21)=城市, W(22)=乡村

    行结构:
      Row 1 = 表头
      Row 2 = CCTV-17 前推期间均值
      Row 3 = CCTV-17 前一天
      Row 4 = CCTV-17 当日
    """
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 4:
        return

    # 找 CCTV-17 行（含"农业农村"）
    cctv17_rows = []
    for r in rows[1:]:
        if r[0] and '农业农村' in str(r[0]):
            cctv17_rows.append(r)
    if len(cctv17_rows) < 2:
        return

    # 下标从 header 取分类名称
    headers = rows[0]
    # H-W = indices 7..22
    CAT_COLS = list(range(7, 23))  # 16 categories
    cat_names = []
    for ci in CAT_COLS:
        h = _safe_str(headers[ci]) if ci < len(headers) and headers[ci] else ''
        # 去掉 "分类观众规模-" 前缀
        h = h.replace('分类观众规模-', '')
        cat_names.append(h)

    period_row = cctv17_rows[0]    # 前推期间均值
    current_row = cctv17_rows[-1]  # 当日

    # 总体观众规模 (G列)
    total_period = _safe_float(period_row[6]) if len(period_row) > 6 else 0
    total_current = _safe_float(current_row[6]) if len(current_row) > 6 else 0

    audience_list = []

    # 总体项
    total_item = AudienceItem()
    total_item.category = '四岁及以上所有人'
    total_item.period_value = total_period
    total_item.current_value = total_current
    if total_period > 0:
        total_item.change = (total_current - total_period) / total_period
    audience_list.append(total_item)

    # 分类项
    for i, ci in enumerate(CAT_COLS):
        name = cat_names[i]
        if not name:
            continue
        pv = _safe_float(period_row[ci]) if ci < len(period_row) else 0
        cv = _safe_float(current_row[ci]) if ci < len(current_row) else 0
        item = AudienceItem()
        item.category = name
        item.period_value = pv
        item.current_value = cv
        if pv > 0:
            item.change = (cv - pv) / pv
        audience_list.append(item)

    if audience_list:
        data.channel_audience = audience_list


def _read_drama(ws, data: ReportData):
    """读取电视剧与非电视剧sheet
    格式:
      Row 0: 日期>>, period_date, period_date, prev_day, prev_day, current, current
      Row 1: 类别, 收视率%, 市场份额%, 收视率%, 市场份额%, 收视率%, 市场份额%
      Row 2: 电视剧, period_r, period_s, prev_r, prev_s, curr_r, curr_s
      Row 3: 非电视剧, ...
    """
    dr = data.drama
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 4:
        return

    # Row 0: 日期标签
    date_row = rows[0]
    dr.period_label = _safe_str(date_row[1])
    n_cols = len([v for v in date_row if v is not None])

    # Row 2: 电视剧
    drama_row = rows[2]
    if drama_row and '电视剧' in _safe_str(drama_row[0]):
        dr.drama_period_rating = _safe_float(drama_row[1])
        dr.drama_period_share = _safe_float(drama_row[2])
        # 7列格式(含前一日): period, prev_day, current
        if len(drama_row) >= 7 and drama_row[5] is not None:
            dr.drama_current_rating = _safe_float(drama_row[5])
            dr.drama_current_share = _safe_float(drama_row[6])
        # 5列格式: period, current
        elif len(drama_row) >= 5:
            dr.drama_current_rating = _safe_float(drama_row[3])
            dr.drama_current_share = _safe_float(drama_row[4])

    # Row 3: 非电视剧
    non_drama_row = rows[3]
    if non_drama_row and ('非电视剧' in _safe_str(non_drama_row[0]) or '其它' in _safe_str(non_drama_row[0])):
        dr.non_drama_period_rating = _safe_float(non_drama_row[1])
        dr.non_drama_period_share = _safe_float(non_drama_row[2])
        if len(non_drama_row) >= 7 and non_drama_row[5] is not None:
            dr.non_drama_current_rating = _safe_float(non_drama_row[5])
            dr.non_drama_current_share = _safe_float(non_drama_row[6])
        elif len(non_drama_row) >= 5:
            dr.non_drama_current_rating = _safe_float(non_drama_row[3])
            dr.non_drama_current_share = _safe_float(non_drama_row[4])


def _read_org_ranking(ws, data: ReportData):
    """读取台组内排名"""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 3:
        return

    for r in rows[2:]:  # 跳过表头(2行)
        if r[0] is None:
            continue
        name = _safe_str(r[0])
        if not name:
            continue
        item = ChannelRankItem()
        item.name = name
        item.prev_share = _safe_float(r[1])
        item.current_share = _safe_float(r[2])
        item.change = item.current_share - item.prev_share
        # 检查是否有简称列
        if len(r) > 3 and r[3]:
            item.short_name = _safe_str(r[3])
        else:
            item.short_name = name
        data.org_ranking.append(item)


def _read_channel_ranking(ws, data: ReportData):
    """读取上星频道排名"""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 3:
        return

    for r in rows[2:]:  # 跳过表头(2行)
        if r[0] is None:
            continue
        name = _safe_str(r[0])
        if not name:
            continue
        item = ChannelRankItem()
        item.name = name
        item.prev_share = _safe_float(r[1])
        item.current_share = _safe_float(r[2])
        if len(r) > 3:
            # 可能有变化列
            if r[3] is not None and isinstance(r[3], (int, float)):
                item.change = _safe_float(r[3])
            elif r[3] is not None:
                item.short_name = _safe_str(r[3])
        if not item.change:
            item.change = item.current_share - item.prev_share
        data.channel_ranking.append(item)


def _read_programs(ws, data: ReportData):
    """读取串单(节目表)"""
    rows = list(ws.iter_rows(min_row=1, values_only=False))
    if len(rows) < 2:
        return

    for row in rows[1:]:  # 跳过表头
        vals = [c.value for c in row]
        if vals[0] is None and vals[2] is None:
            continue
        item = ProgramItem()
        item.name = _safe_str(vals[0])
        item.episode = _safe_str(vals[1])
        item.date = _date_to_str(vals[3])
        item.weekday = _safe_str(vals[4])
        item.start_time = _time_to_str(vals[5])
        item.duration = int(_safe_float(vals[6]))
        item.end_time = _time_to_str(vals[7])
        item.category = _safe_str(vals[9])
        item.subcategory = _safe_str(vals[10])
        item.market_share = _safe_float(vals[11])
        item.rating = _safe_float(vals[12])
        data.programs.append(item)


def _read_minutes(ws, data: ReportData):
    """读取分分钟数据"""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 3:
        return

    for r in rows[2:]:  # 跳过表头(2行)
        if r[0] is None:
            continue
        item = MinuteData()
        item.time_str = _time_to_str(r[0])
        item.period_rating = _safe_float(r[1])
        item.current_rating = _safe_float(r[2])
        item.period_share = _safe_float(r[3])
        item.current_share = _safe_float(r[4])
        data.minutes.append(item)


def _read_time_slots(ws, data: ReportData):
    """读取日报-分时段收视率"""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 3:
        return

    for r in rows[2:]:  # 跳过表头
        if r[0] is None:
            continue
        item = TimeSlotData()
        item.slot_name = _safe_str(r[0])
        item.period_rating = _safe_float(r[1])
        item.current_rating = _safe_float(r[2])
        item.period_share = _safe_float(r[3])
        item.current_share = _safe_float(r[4])
        item.rating_change = _safe_float(r[5])
        item.share_change = _safe_float(r[6])
        data.time_slots.append(item)


def _read_theaters(ws, data: ReportData):
    """读取日报-三大剧场"""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 3:
        return

    for r in rows[2:]:
        if r[0] is None and r[1] is None:
            continue
        item = TheaterData()
        item.time_slot = _safe_str(r[0])
        item.info = _safe_str(r[1])
        item.period_rating = _safe_float(r[2])
        item.period_share = _safe_float(r[3])
        item.current_rating = _safe_float(r[4])
        item.current_share = _safe_float(r[5])
        item.rating_change = _safe_float(r[6])
        item.share_change = _safe_float(r[7])
        data.theaters.append(item)


def _read_premiere_programs(ws, data: ReportData):
    """读取日报-首播节目对比"""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        return

    for r in rows[1:]:
        if r[0] is None:
            continue
        item = PremiereProgram()
        item.name = _safe_str(r[0])
        item.rating = _safe_float(r[4])
        item.share = _safe_float(r[5])
        item.prev_day_rating = _safe_float(r[6])
        item.prev_day_share = _safe_float(r[7])
        item.period_rating = _safe_float(r[8])
        item.period_share = _safe_float(r[9])
        item.rating_change = _safe_float(r[10])
        item.share_change = _safe_float(r[11])
        if len(r) > 12 and r[12]:
            item.time_slot = _safe_str(r[12])
        data.premiere_programs.append(item)


def _find_premiere_col(header_row):
    """在串单表头中查找"首播"标注列。

    扫描表头行，找到包含"首播"文字的列索引。
    返回列索引(int)或 None（未找到）。
    """
    for i, v in enumerate(header_row):
        if v is not None and '首播' in _safe_str(v):
            return i
    return None


def _read_premiere_from_schedule(ws, data: ReportData):
    """从串单sheet读取首播节目数据。

    筛选策略:
      - 优先: 若串单最后几列中有"首播"标注列（表头含"首播"），
        则只选取该列标记了"首播"的行（用户手动标注）。
      - 兜底: 若无"首播"标注列，使用自动筛选（16:30-24:00 非电视剧）。

    列映射:
      A(0)=名称, F(5)=开始时间, H(7)=结束时间,
      J(9)=类别, L(11)=市场份额%, M(12)=收视率%,
      N(13)=前一天同时段收视率%, O(14)=前一天同时段市场份额%,
      P(15)=前1个月同时段收视率%(可选), Q(16)=前1个月同时段市场份额%(可选)
    """
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        return

    import datetime as _dt

    # 检测是否有"首播"标注列
    premiere_col = _find_premiere_col(rows[0])
    use_manual_filter = premiere_col is not None
    if use_manual_filter:
        print(f'  串单: 检测到"首播"标注列(第{premiere_col+1}列)，按标注筛选首播节目')

    cutoff_start = _dt.time(16, 30, 0)

    premiere_list = []
    for r in rows[1:]:
        if r[0] is None:
            continue

        # ── 筛选逻辑 ──
        if use_manual_filter:
            # 手动标注模式: 只取标注了"首播"的行
            mark = _safe_str(r[premiere_col]) if len(r) > premiere_col else ''
            if '首播' not in mark:
                continue
        else:
            # 自动筛选模式: 16:30起 + 排除电视剧
            raw_start = r[5]
            if raw_start is None:
                continue
            if isinstance(raw_start, _dt.datetime):
                start_t = raw_start.time()
            elif isinstance(raw_start, _dt.time):
                start_t = raw_start
            else:
                continue
            if start_t < cutoff_start:
                continue
            cat = _safe_str(r[9]) if len(r) > 9 else ''
            if '电视剧' in cat:
                continue

        # ── 解析开始/结束时间 ──
        raw_start = r[5]
        if raw_start is None:
            continue
        if isinstance(raw_start, _dt.datetime):
            start_t = raw_start.time()
        elif isinstance(raw_start, _dt.time):
            start_t = raw_start
        else:
            continue
        start_str = start_t.strftime('%H:%M')

        raw_end = r[7] if len(r) > 7 else None
        if isinstance(raw_end, _dt.datetime):
            end_str = raw_end.strftime('%H:%M')
        elif isinstance(raw_end, _dt.time):
            end_str = raw_end.strftime('%H:%M')
        else:
            end_str = _time_to_str(raw_end)

        item = PremiereProgram()
        item.name = _safe_str(r[0])
        item.time_slot = f'{start_str}-{end_str}'
        item.rating = _safe_float(r[12]) if len(r) > 12 else 0
        item.share = _safe_float(r[11]) if len(r) > 11 else 0
        item.prev_day_rating = _safe_float(r[13]) if len(r) > 13 else 0
        item.prev_day_share = _safe_float(r[14]) if len(r) > 14 else 0
        # 前1个月数据（P、Q列，可选）
        item.period_rating = _safe_float(r[15]) if len(r) > 15 else 0
        item.period_share = _safe_float(r[16]) if len(r) > 16 else 0
        # 计算变化（当日 vs 前一天）
        if item.prev_day_rating and item.prev_day_rating > 0:
            item.rating_change = round((item.rating - item.prev_day_rating) / item.prev_day_rating * 100)
        if item.prev_day_share and item.prev_day_share > 0:
            item.share_change = round((item.share - item.prev_day_share) / item.prev_day_share * 100)

        premiere_list.append(item)

    # 用串单数据替换旧的 premiere_programs
    if premiere_list:
        data.premiere_programs = premiere_list
        if use_manual_filter:
            print(f'  串单: 共筛选出 {len(premiere_list)} 个首播节目')


def _read_premiere_audience(ws, data: ReportData):
    """读取日报-首播节目分类观众触达"""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 3:
        return

    # 第1行: 节目名称
    header = rows[0]
    programs = []
    for v in header[1:]:
        if v is not None:
            programs.append(_safe_str(v))

    # 第3行开始: 分类观众数据
    categories = {}
    for r in rows[2:]:
        if r[0] is None:
            continue
        cat = _safe_str(r[0])
        if not cat:
            continue
        vals = [_safe_float(v) for v in r[1:len(programs)+1]]
        categories[cat] = vals

    data.premiere_audience = {
        'programs': programs,
        'categories': categories
    }

    # 读取首播节目观众总触达
    for r in rows:
        if r and r[0] and '首播节目观众总触达' in str(r[0]):
            data.premiere_audience['total_reach'] = _safe_float(r[1])
            data.premiere_audience['period_avg_reach'] = _safe_float(r[2]) if len(r) > 2 else 0
            data.premiere_audience['reach_change'] = _safe_float(r[3]) if len(r) > 3 else 0


def _read_channel_audience(ws, data: ReportData):
    """读取日报-频道分类观众规模"""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 4:
        return

    # 使用右侧万人数据 (G-J列)
    for r in rows[3:]:  # 跳过3行表头
        if r[6] is None:
            continue
        item = AudienceItem()
        item.category = _safe_str(r[6])
        if not item.category:
            continue
        item.period_value = _safe_float(r[7])
        item.current_value = _safe_float(r[8])
        if len(r) > 9 and r[9] is not None:
            item.change = _safe_float(r[9])
        elif item.period_value > 0:
            item.change = (item.current_value - item.period_value) / item.period_value
        data.channel_audience.append(item)


def _read_drama_audience(ws, data: ReportData):
    """读取日报-电视剧观众规模"""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 3:
        return

    # 左侧: 电视剧分类观众(B-F列)
    audience_data = {}
    current_section = None

    for r in rows[2:]:  # 跳过2行表头
        if r[1] is None and r[2] is None:
            continue
        time_slot = _safe_str(r[1])
        # 时间段格式如 "12:20 - 18:00"
        if time_slot and ':' in time_slot and any(c.isdigit() for c in time_slot):
            current_section = time_slot

        target = _safe_str(r[2])
        if not target or target in ('目标', '摘要'):
            continue

        if current_section is None:
            continue

        if current_section not in audience_data:
            audience_data[current_section] = []
        audience_data[current_section].append({
            'target': target,
            'rating': _safe_float(r[3]),
            'share': _safe_float(r[4]),
            'audience': _safe_float(r[5]),  # 观众规模(万人)
        })

    # 右侧: 各剧场观众规模 (H-K列)
    theater_audience = []
    for r in rows[2:]:
        if r is None or len(r) < 11:
            continue
        name = _safe_str(r[7])
        if not name or name in ('', '名称', '剧场'):
            continue
        if r[9] is not None:
            theater_audience.append({
                'name': name,
                'time_slot': _safe_str(r[8]),
                'period_value': _safe_float(r[9]),
                'current_value': _safe_float(r[10]),
            })

    data.drama_audience = {
        'by_section': audience_data,
        'theaters': theater_audience
    }


def _compute_rankings(data: ReportData):
    """计算CCTV-17在各排名中的位置"""
    # 台组内排名
    for i, item in enumerate(data.org_ranking):
        if '农业农村' in item.name or 'CCTV-17' in item.name or 'CCTV-17' in item.short_name:
            data.org_rank = i  # 注意第一行可能是台组合计
            # 计算变化: 根据当日与前日的排名差
            prev_sorted = sorted([x for x in data.org_ranking if '中央级' not in x.name],
                               key=lambda x: x.prev_share, reverse=True)
            curr_sorted = sorted([x for x in data.org_ranking if '中央级' not in x.name],
                               key=lambda x: x.current_share, reverse=True)
            prev_rank = next((i+1 for i, x in enumerate(prev_sorted)
                            if '农业农村' in x.name or 'CCTV-17' in x.short_name), 0)
            curr_rank = next((i+1 for i, x in enumerate(curr_sorted)
                            if '农业农村' in x.name or 'CCTV-17' in x.short_name), 0)
            data.org_rank = curr_rank
            data.org_rank_change = prev_rank - curr_rank  # 正数=上升
            break

    # 上星频道排名
    curr_sorted = sorted(data.channel_ranking, key=lambda x: x.current_share, reverse=True)
    prev_sorted = sorted(data.channel_ranking, key=lambda x: x.prev_share, reverse=True)
    for i, item in enumerate(curr_sorted):
        name = item.name
        if '农业农村' in name or 'CCTV-17' in name:
            data.channel_rank = i + 1
            prev_rank = next((j+1 for j, x in enumerate(prev_sorted)
                            if '农业农村' in x.name or 'CCTV-17' in x.name), 0)
            data.channel_rank_change = prev_rank - data.channel_rank
            break


def _compute_premiere_loyalty(data: ReportData):
    """
    计算首播节目忠实度。
    忠实度 = 收视率(%) / 到达率(%) × 100
    到达率(%) = 观众规模(万人) / TV总人口(万人) × 100
    TV总人口 = 频道到达率(万人) / (频道到达率% / 100)
    """
    if not data.premiere_programs or not data.premiere_audience:
        return

    ms = data.market_share
    # 从频道观众规模推导TV总人口(万人)
    # 从频道分类观众规模中: 频道到达率(万人)
    channel_reach_wan = 0
    for a in data.channel_audience:
        if '四岁' in a.category and '以上' in a.category:
            channel_reach_wan = a.current_value
            break

    # 频道到达率(%): 来自市场份额
    channel_reach_pct = ms.cctv17_current_reach  # e.g. 1.559

    if channel_reach_pct <= 0 or channel_reach_wan <= 0:
        return

    # TV总人口(万人) = 到达率(万人) / (到达率(%)/100)
    universe = channel_reach_wan / (channel_reach_pct / 100.0)

    # 获取首播节目观众数据 (四岁以上)
    progs = data.premiere_audience.get('programs', [])
    cats = data.premiere_audience.get('categories', {})
    audience_row = cats.get('四岁及以上所有人', [])

    # 按节目名称建立观众规模映射
    audience_map = {}
    for i, pname in enumerate(progs):
        if i < len(audience_row):
            audience_map[pname] = audience_row[i]

    # 为每个首播节目计算忠实度
    for p in data.premiere_programs:
        aud = audience_map.get(p.name, 0)
        p.audience = aud
        if aud > 0 and universe > 0:
            reach_pct = aud / universe * 100.0
            p.loyalty = p.rating / reach_pct * 100.0 if reach_pct > 0 else 0
